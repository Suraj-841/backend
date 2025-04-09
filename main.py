from fastapi import FastAPI, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import os
import openpyxl
from excel_utils import (
    load_excel,
    save_excel,
    run_daily_check,
    get_all_students,
    get_expired_students_data,
    update_expiry_in_excel,
    replace_student_in_excel
)
from notifier import send_push_notification  # <== Add this line

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class StatusUpdateRequest(BaseModel):
    seat_no: int
    new_status: str


class UpdateExpiryRequest(BaseModel):
    seat_no: int
    name: str
    new_expiry: str

class ReplaceStudentRequest(BaseModel):
    seat_no: int
    name: str
    day_type: str
    charge: int
    start_date: str
    phone: Optional[str] = ""
    status: str

@app.get("/students")
def get_students():
    return get_all_students()

@app.get("/expired-students")
def get_expired_students():
    return get_expired_students_data()

@app.post("/update-expiry")
def update_expiry(req: UpdateExpiryRequest):
    updated = update_expiry_in_excel(req)
    if not updated:
        raise HTTPException(status_code=404, detail="Student not found")

    send_push_notification(f"📆 Expiry updated for {req.name} (Seat {req.seat_no}) to {req.new_expiry}.")
    return {"message": "Expiry updated successfully."}

@app.post("/replace-student")
def replace_student(req: ReplaceStudentRequest):
    result = replace_student_in_excel(req)
    if not result:
        raise HTTPException(status_code=400, detail="Failed to replace student")

    if req.name.lower() == "vacant":
        send_push_notification(f"🪑 Seat {req.seat_no} has been vacated.")
    else:
        send_push_notification(f"👤 {req.name} has been assigned to Seat {req.seat_no}.")

    return {"message": "Student replaced successfully."}


@app.post("/update-status")
def update_status(req: StatusUpdateRequest):
    wb, sheet = load_excel()
    found = False

    for row in range(2, sheet.max_row + 1):
        if sheet[f"A{row}"].value == req.seat_no:
            sheet[f"G{row}"] = req.new_status
            found = True
            break

    if not found:
        raise HTTPException(status_code=404, detail="Seat not found")

    save_excel(wb)
    send_push_notification(f"💰 Seat {req.seat_no} status updated to {req.new_status}")
    return {"message": "Status updated successfully."}


@app.get("/")
def root():
    return {"message": "Backend running ✅"}

@app.get("/left-students")
def get_left_students():
    filename = "Left_Students_Log.xlsx"
    if not os.path.exists(filename):
        return []
    
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    data = []
    for row in range(2, sheet.max_row + 1):
        student = {
            "Seat No": sheet[f"A{row}"].value,
            "Name": sheet[f"B{row}"].value,
            "Phone": sheet[f"C{row}"].value,
            "Start Date": sheet[f"D{row}"].value,
            "Left On": sheet[f"E{row}"].value,
            "Status": sheet[f"F{row}"].value,
            "Reason": sheet[f"G{row}"].value,
        }
        data.append(student)
    return data




@app.get("/daily-check")
def daily_check():
    result = run_daily_check()
    if result["count"] > 0:
        send_push_notification(f"🚨 {result['count']} student(s) expired. Daily check complete.")
    else:
        send_push_notification("✅ Daily check complete. No expired students today.")
    return result
