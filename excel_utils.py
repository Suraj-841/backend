# === excel_utils.py (Modular Excel Utilities with Logging for Left Students) ===

import openpyxl
from datetime import datetime, timedelta
from notifier import send_push_notification
import os
from dateutil import parser

EXCEL_FILE = "Student_Seat_Assignment.xlsx"
LEFT_LOG_FILE = "Left_Students_Log.xlsx"

def load_excel():
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"{EXCEL_FILE} not found.")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    return wb, sheet

def save_excel(wb):
    wb.save(EXCEL_FILE)

def get_all_students():
    wb, sheet = load_excel()
    data = []
    for row in range(2, sheet.max_row + 1):
        student = {
            "Seat No": sheet[f"A{row}"].value,
            "Name": sheet[f"B{row}"].value,
            "Day Type": sheet[f"C{row}"].value,
            "Charge": sheet[f"D{row}"].value,
            "Start Date": sheet[f"E{row}"].value,
            "Expiry Date": sheet[f"F{row}"].value,
            "Status": sheet[f"G{row}"].value,
            "Phone": sheet[f"H{row}"].value,
        }
        data.append(student)
    return data

def get_expired_students_data():
    wb, sheet = load_excel()
    today = datetime.today().date()
    expired_list = []

    for row in range(2, sheet.max_row + 1):
        expiry = sheet[f"F{row}"].value
        expiry_date = None

        # Try parsing manually if it's not a datetime
        if isinstance(expiry, datetime):
            expiry_date = expiry.date()
        elif isinstance(expiry, str):
            try:
                expiry_date = parser.parse(expiry).date()
            except:
                continue

        if expiry_date and expiry_date < today:
            student = {
                "Seat No": sheet[f"A{row}"].value,
                "Name": sheet[f"B{row}"].value,
                "Day Type": sheet[f"C{row}"].value,
                "Charge": sheet[f"D{row}"].value,
                "Start Date": sheet[f"E{row}"].value,
                "Expiry Date": expiry,
                "Status": sheet[f"G{row}"].value,
                "Phone": sheet[f"H{row}"].value,
            }
            expired_list.append(student)

    return expired_list

def update_expiry_in_excel(req):
    wb, sheet = load_excel()
    for row in range(2, sheet.max_row + 1):
        if sheet[f"A{row}"].value == req.seat_no and sheet[f"B{row}"].value == req.name:
            sheet[f"F{row}"] = req.new_expiry
            break
    save_excel(wb)
    return {"message": "Expiry updated successfully."}

def log_left_student(data):
    if not os.path.exists(LEFT_LOG_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Seat No", "Name", "Phone", "Start Date", "Left On", "Status", "Reason"])
    else:
        wb = openpyxl.load_workbook(LEFT_LOG_FILE)
        sheet = wb.active

    today = datetime.today().strftime("%d %B %Y")
    sheet.append([
        data.get("seat_no"),
        data.get("name"),
        data.get("phone", ""),
        data.get("start_date", ""),
        today,
        data.get("status", ""),
        "Vacated"
    ])
    wb.save(LEFT_LOG_FILE)

def replace_student_in_excel(req):
    wb, sheet = load_excel()
    for row in range(2, sheet.max_row + 1):
        if sheet[f"A{row}"].value == req.seat_no:

            # ✅ If student is being vacated — log their details
            if req.name.lower() == "vacant":
                log_left_student({
                    "seat_no": sheet[f"A{row}"].value,
                    "name": sheet[f"B{row}"].value,
                    "phone": sheet[f"H{row}"].value,
                    "start_date": sheet[f"E{row}"].value,
                    "status": sheet[f"G{row}"].value
                })

            # ✅ Replace student (new or vacant)
            sheet[f"B{row}"] = req.name
            sheet[f"C{row}"] = req.day_type
            sheet[f"D{row}"] = req.charge
            sheet[f"E{row}"] = req.start_date
            sheet[f"F{row}"] = ""
            sheet[f"G{row}"] = req.status
            sheet[f"H{row}"] = req.phone

            try:
                start = datetime.strptime(req.start_date, "%d %B")
                start = start.replace(year=datetime.now().year)
                expiry = start + timedelta(days=30)
                sheet[f"F{row}"] = expiry.strftime("%d %B %Y")
            except Exception as e:
                print("Date parsing error:", e)

            break
    save_excel(wb)
    return {"message": "Student replaced successfully."}

def run_daily_check():
    wb, sheet = load_excel()
    today = datetime.today().date()
    expired = []
    for row in range(2, sheet.max_row + 1):
        name = sheet[f"B{row}"].value
        expiry = sheet[f"F{row}"].value
        status = sheet[f"G{row}"].value
        seat = sheet[f"A{row}"].value
        if name and expiry and isinstance(expiry, datetime):
            if expiry.date() < today and status != "Pending":
                sheet[f"G{row}"] = "Pending"
                msg = f"{name} (Seat {seat}) membership expired."
                send_push_notification(msg)
                expired.append(msg)
    save_excel(wb)
    return {"expired_students": expired, "count": len(expired)}
